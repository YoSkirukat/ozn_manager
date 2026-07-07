"""Экспорт заказов в Excel."""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.datetime_fmt import format_datetime, utc_bounds_for_local_dates
from app.models import HOME_COUNTRY_NAME, ORDER_STATUS_DELIVERED, Order, User
from app.services.order_details import (
    _apply_product_margins,
    _cached_total_accrued,
    _product_lookup,
    _product_rows,
    attach_order_margins,
)
from app.services.order_returns import build_post_delivery_return_postings

EXCEL_HEADERS = (
    "Дата",
    "Номер заказа",
    "Баркод",
    "Статус",
    "Наименование",
    "Кол-во",
    "Сумма",
    "Начислено",
    "Себестоимость",
    "Маржа",
    "Схема работы",
    "Регион",
)

ONEC_HEADERS = (
    "Баркод",
    "Кол-во",
    "Цена",
    "Сумма",
)


def fetch_orders_for_export(
    user: User,
    date_from: date,
    date_to: date,
    statuses: list[str] | None = None,
    schemes: list[str] | None = None,
    delivery: str = "",
) -> list[Order]:
    start, end = utc_bounds_for_local_dates(date_from, date_to)
    query = Order.query.filter(
        Order.user_id == user.id,
        Order.order_date >= start,
        Order.order_date <= end,
    )
    if statuses:
        query = query.filter(Order.status.in_(statuses))
    if schemes:
        query = query.filter(Order.scheme.in_(schemes))
    orders = query.order_by(Order.order_date.desc(), Order.id.desc()).all()
    from app.services.orders_filters import apply_delivery_filter

    orders = apply_delivery_filter(orders, delivery)
    attach_order_margins(orders, user, use_transactions=False)
    return orders


def _decimal(value) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal(0)


def _money(value) -> float | None:
    if value is None:
        return None
    return float(_decimal(value).quantize(Decimal("0.01")))


def _order_total_accrued(order: Order) -> float | None:
    raw = order.raw_data if isinstance(order.raw_data, dict) else {}
    cached = _cached_total_accrued(raw)
    return cached


def _region_label(order: Order) -> str:
    return order.delivery_country() or HOME_COUNTRY_NAME


def build_excel_export_rows(orders: list[Order], user: User) -> list[tuple]:
    product_lookup = _product_lookup(user.id)
    rows: list[tuple] = []

    for order in orders:
        raw = order.raw_data if isinstance(order.raw_data, dict) else {}
        total_accrued = _order_total_accrued(order)
        products = _product_rows(
            user.id,
            raw,
            order.thumbnail_url,
            product_lookup=product_lookup,
        )
        products = _apply_product_margins(
            products,
            total_accrued or 0,
            is_international=order.is_international(),
            calculate=order.status == ORDER_STATUS_DELIVERED,
        )

        sale_amounts = [_decimal(p.get("sale_amount")) for p in products]
        sale_sum = sum(sale_amounts) or Decimal(0)
        region = _region_label(order)
        order_date = format_datetime(order.order_date, "%d.%m.%Y %H:%M")

        if not products:
            rows.append(
                (
                    order_date,
                    order.ozon_order_id,
                    "—",
                    order.status_display(),
                    "—",
                    None,
                    _money(order.total),
                    total_accrued,
                    None,
                    order.order_margin(user) if order.status == ORDER_STATUS_DELIVERED else None,
                    order.scheme_display(),
                    region,
                )
            )
            continue

        for product, sale_amount in zip(products, sale_amounts):
            qty = float(product.get("quantity") or 1)
            purchase_unit = product.get("purchase_price")
            cost_total = (
                _money(_decimal(purchase_unit) * _decimal(qty))
                if purchase_unit is not None
                else None
            )
            if total_accrued is not None and sale_sum != 0:
                accrued_share = _money(_decimal(total_accrued) * (sale_amount / sale_sum))
            else:
                accrued_share = None

            rows.append(
                (
                    order_date,
                    order.ozon_order_id,
                    product.get("barcode") or "—",
                    order.status_display(),
                    product.get("name") or "—",
                    qty,
                    _money(sale_amount),
                    accrued_share,
                    cost_total,
                    product.get("margin"),
                    order.scheme_display(),
                    region,
                )
            )

    return rows


def build_onec_export_rows(orders: list[Order], user: User) -> list[tuple]:
    product_lookup = _product_lookup(user.id)
    # Для 1С важно, чтобы "Сумма" соответствовала "Начислено", а не выручке.
    # Поэтому суммируем долю начисленного из общего total_accrued заказа.
    aggregated: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"qty": Decimal(0), "sum": Decimal(0)}
    )
    returned_postings = build_post_delivery_return_postings(user.id)

    for order in orders:
        if order.status != ORDER_STATUS_DELIVERED:
            continue
        if order.ozon_order_id in returned_postings:
            continue
        raw = order.raw_data if isinstance(order.raw_data, dict) else {}
        total_accrued = _order_total_accrued(order)
        products = _product_rows(
            user.id,
            raw,
            order.thumbnail_url,
            product_lookup=product_lookup,
        )
        sale_amounts = [_decimal(p.get("sale_amount")) for p in products]
        sale_sum = sum(sale_amounts) or Decimal(0)
        for product in products:
            barcode = str(product.get("barcode") or "").strip()
            if not barcode or barcode == "—":
                continue
            qty = _decimal(product.get("quantity") or 1)
            sale_amount = _decimal(product.get("sale_amount"))
            aggregated[barcode]["qty"] = _decimal(aggregated[barcode]["qty"]) + qty
            if total_accrued is not None and sale_sum != 0:
                accrued_share = _decimal(total_accrued) * (sale_amount / sale_sum)
            else:
                # Если total_accrued отсутствует (нет кэша/данных), оставляем прежнее поведение,
                # чтобы выгрузка не превращалась в нули.
                accrued_share = sale_amount
            aggregated[barcode]["sum"] = _decimal(aggregated[barcode]["sum"]) + accrued_share

    rows: list[tuple] = []
    for barcode in sorted(aggregated):
        data = aggregated[barcode]
        qty = float(data["qty"])
        total_sum = _decimal(data["sum"])
        price = _money(total_sum / _decimal(qty)) if qty else 0.0
        rows.append((barcode, qty, price, _money(total_sum)))
    return rows


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 48)


def build_xlsx(headers: tuple[str, ...], rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    header_font = Font(bold=True)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    _autosize_columns(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_orders_excel(
    user: User,
    date_from: date,
    date_to: date,
    *,
    export_type: str,
    statuses: list[str] | None = None,
    schemes: list[str] | None = None,
    delivery: str = "",
) -> tuple[bytes, str]:
    orders = fetch_orders_for_export(
        user,
        date_from,
        date_to,
        statuses,
        schemes,
        delivery,
    )
    period = f"{date_from.isoformat()}_{date_to.isoformat()}"

    if export_type == "1c":
        # Для 1С лучше сразу гарантировать наличие `_total_accrued` в `raw_data`,
        # иначе придётся падать на "выручку" (sale_amount), что неверно для "Начислено".
        attach_order_margins(orders, user, use_transactions=True)
        rows = build_onec_export_rows(orders, user)
        content = build_xlsx(ONEC_HEADERS, rows)
        filename = f"orders_1c_{period}.xlsx"
        return content, filename

    rows = build_excel_export_rows(orders, user)
    content = build_xlsx(EXCEL_HEADERS, rows)
    filename = f"orders_{period}.xlsx"
    return content, filename
