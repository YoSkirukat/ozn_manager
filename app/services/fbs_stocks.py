"""Остатки FBS из внешнего Excel + выгрузка в Ozon."""

from __future__ import annotations

import io
import re

import requests

from app.db_sqlite import db_session_commit
from app.models import Product
from app.ozon.fbs_stocks import resolve_fbs_warehouse_id, update_fbs_stocks
from app.services.purchase_prices import normalize_barcode, normalize_sheet_url

BARCODE_HEADERS = {"баркод", "barcode", "штрихкод", "штрих-код"}
STOCK_HEADERS = {
    "остаток fbs",
    "остатки fbs",
    "остаток",
    "остатки",
    "количество",
    "кол-во",
    "stock",
    "stock_fbs",
    "fbs",
}


def _normalize_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _parse_stock(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return max(0, value)
    if isinstance(value, float):
        if value != value:
            return None
        return max(0, int(value))
    text = str(value).strip().replace(" ", "").replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text:
        return None
    try:
        return max(0, int(float(text)))
    except (TypeError, ValueError):
        return None


def _find_columns(headers: list) -> tuple[int | None, int | None]:
    barcode_col = None
    stock_col = None
    for idx, header in enumerate(headers):
        norm = _normalize_header(header)
        if norm in BARCODE_HEADERS:
            barcode_col = idx
        if norm in STOCK_HEADERS or norm.startswith("остаток"):
            stock_col = idx
    return barcode_col, stock_col


def _parse_workbook(content: bytes) -> dict[str, int]:
    if content[:2] == b"PK":
        return _parse_xlsx(content)
    return _parse_xls(content)


def _parse_xlsx(content: bytes) -> dict[str, int]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _rows_to_stocks(rows)


def _parse_xls(content: bytes) -> dict[str, int]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(r) for r in range(sheet.nrows)]
    return _rows_to_stocks(rows)


def _rows_to_stocks(rows: list) -> dict[str, int]:
    if not rows:
        raise ValueError("Файл пуст.")

    header_row = None
    header_index = 0
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        bc, st = _find_columns(list(row))
        if bc is not None and st is not None:
            header_row = list(row)
            header_index = i
            break

    if header_row is None:
        raise ValueError('Не найдены колонки «Баркод» и «Остаток FBS».')

    barcode_col, stock_col = _find_columns(header_row)
    stocks: dict[str, int] = {}

    for row in rows[header_index + 1 :]:
        if not row or len(row) <= max(barcode_col, stock_col):
            continue
        barcode = normalize_barcode(row[barcode_col])
        stock = _parse_stock(row[stock_col])
        if barcode and stock is not None:
            stocks[barcode] = stock

    if not stocks:
        raise ValueError("В файле нет строк с баркодом и остатком.")
    return stocks


def fetch_fbs_stocks_map(url: str) -> dict[str, int]:
    download_url = normalize_sheet_url(url)
    try:
        resp = requests.get(download_url, timeout=60)
    except requests.RequestException as exc:
        raise RuntimeError(f"Не удалось скачать файл: {exc}") from exc
    if resp.status_code != 200:
        raise RuntimeError(f"Ошибка загрузки файла: HTTP {resp.status_code}")
    return _parse_workbook(resp.content)


def _preferred_warehouse_id(user) -> int | None:
    raw = (getattr(user, "fbs_warehouse_id", None) or "").strip()
    if not raw:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _build_ozon_items(user, stock_map: dict[str, int]) -> tuple[list[dict], int]:
    """Позиции с изменившимся остатком + число совпадений по баркоду."""
    products = Product.query.filter_by(user_id=user.id).all()
    items: list[dict] = []
    matched = 0

    for product in products:
        barcode = normalize_barcode(product.barcode)
        if not barcode:
            continue
        stock = stock_map.get(barcode)
        if stock is None:
            continue
        matched += 1

        if int(product.stock_fbs or 0) == stock:
            continue

        entry: dict = {
            "stock": stock,
            "product": product,
        }
        if product.ozon_product_id:
            entry["product_id"] = product.ozon_product_id
        if product.offer_id:
            entry["offer_id"] = product.offer_id
        if "product_id" not in entry and "offer_id" not in entry:
            continue
        items.append(entry)

    return items, matched


def _apply_local_stock_updates(updated_items: list[dict]) -> int:
    local_updated = 0
    for item in updated_items:
        product = item.get("product")
        if product is None:
            continue
        stock = int(item["stock"])
        if int(product.stock_fbs or 0) != stock:
            local_updated += 1
        product.stock_fbs = stock
    return local_updated


def _item_lookup_key(item: dict) -> str:
    if item.get("offer_id") not in (None, ""):
        return f"offer:{item['offer_id']}"
    if item.get("product_id") not in (None, ""):
        return f"product:{item['product_id']}"
    return ""


def _index_items_by_keys(items: list[dict]) -> dict[str, dict]:
    by_key: dict[str, dict] = {}
    for item in items:
        if item.get("offer_id") not in (None, ""):
            by_key[f"offer:{item['offer_id']}"] = item
        product_id = item.get("product_id")
        if product_id not in (None, ""):
            by_key[f"product:{product_id}"] = item
            try:
                by_key[f"product:{int(product_id)}"] = item
            except (TypeError, ValueError):
                pass
    return by_key


def _apply_stock_map(user, stock_map: dict[str, int], *, push_to_ozon: bool = True) -> dict:
    items, matched_in_file = _build_ozon_items(user, stock_map)

    result = {
        "ok": True,
        "skipped": False,
        "updated": 0,
        "total_in_file": len(stock_map),
        "matched": matched_in_file,
        "changed": len(items),
        "ozon_updated": 0,
        "ozon_failed": 0,
        "ozon_deferred": 0,
    }

    if not push_to_ozon:
        local_updated = _apply_local_stock_updates(items)
        result["updated"] = local_updated
        result["message"] = (
            f"Остатки FBS в приложении: обновлено {local_updated} "
            f"из {len(stock_map)} в файле."
        )
        return result

    if not user.has_ozon_credentials():
        result["ok"] = False
        result["error"] = "Подключите Ozon API в профиле, чтобы выгрузить остатки в кабинет."
        result["message"] = "В Ozon не отправлено: нет ключей API."
        return result

    if matched_in_file == 0:
        result["message"] = (
            f"В файле {len(stock_map)} строк, но нет совпадений с товарами в каталоге "
            "(по баркоду)."
        )
        return result

    if not items:
        result["message"] = (
            f"В файле {len(stock_map)} строк, совпадений по баркоду {matched_in_file}. "
            "Изменений остатков нет — в Ozon ничего не отправляли."
        )
        return result

    by_key = _index_items_by_keys(items)
    ozon_payload = [
        {k: v for k, v in item.items() if k != "product"}
        for item in items
    ]

    try:
        warehouse_id = resolve_fbs_warehouse_id(
            user.ozon_client_id,
            user.ozon_api_key,
            _preferred_warehouse_id(user),
        )
        ozon_result = update_fbs_stocks(
            user.ozon_client_id,
            user.ozon_api_key,
            warehouse_id,
            ozon_payload,
        )
    except Exception as exc:
        result["ok"] = False
        result["error"] = str(exc)
        result["message"] = f"Ошибка выгрузки в Ozon: {exc}"
        return result

    updated_with_products: list[dict] = []
    for raw in ozon_result.get("updated_items") or []:
        src = by_key.get(_item_lookup_key(raw))
        if src is None and raw.get("product_id") not in (None, ""):
            src = by_key.get(f"product:{raw['product_id']}")
        if src is not None:
            updated_with_products.append(src)

    local_updated = _apply_local_stock_updates(updated_with_products)
    result["updated"] = local_updated
    result["ozon_updated"] = ozon_result.get("updated", 0)
    result["ozon_failed"] = ozon_result.get("failed", 0)
    result["ozon_deferred"] = ozon_result.get("deferred", 0)
    result["warehouse_id"] = warehouse_id

    deferred = result["ozon_deferred"]
    if ozon_result.get("ok"):
        if deferred:
            result["message"] = (
                f"Остатки FBS: в Ozon обновлено {result['ozon_updated']}, "
                f"отложено из‑за частоты {deferred} (склад {warehouse_id})."
            )
        else:
            result["message"] = (
                f"Остатки FBS: локально {local_updated}, в Ozon обновлено "
                f"{result['ozon_updated']} (склад {warehouse_id})."
            )
    else:
        result["ok"] = False
        err_tail = ""
        errors = ozon_result.get("errors") or []
        if errors:
            err_tail = " " + "; ".join(errors[:5])
        result["error"] = (
            f"Часть остатков не обновлена в Ozon: успешно {result['ozon_updated']}, "
            f"ошибок {result['ozon_failed']}.{err_tail}"
        )
        result["message"] = (
            f"Локально обновлено {local_updated}. "
            f"В Ozon: успешно {result['ozon_updated']}, ошибок {result['ozon_failed']}."
            f"{err_tail}"
        )

    return result


def apply_fbs_stocks_from_content(user, content: bytes) -> dict:
    try:
        stock_map = _parse_workbook(content)
    except Exception as exc:
        return {"ok": False, "error": str(exc), "updated": 0}
    return _apply_stock_map(user, stock_map, push_to_ozon=True)


def apply_fbs_stocks(user) -> dict:
    url = (getattr(user, "fbs_stocks_url", None) or "").strip()
    if not url:
        return {
            "ok": False,
            "skipped": True,
            "updated": 0,
            "error": "Укажите ссылку на файл «Остатки для FBS» в профиле.",
        }

    try:
        stock_map = fetch_fbs_stocks_map(url)
    except Exception as exc:
        return {"ok": False, "error": str(exc), "updated": 0}

    result = _apply_stock_map(user, stock_map, push_to_ozon=True)
    db_session_commit()
    return result
