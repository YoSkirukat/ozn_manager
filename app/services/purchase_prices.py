"""Загрузка закупочных цен из XLS/XLSX по ссылке."""

import io
import re
from decimal import Decimal, InvalidOperation
from urllib.parse import parse_qs, urlparse

import requests

from app.extensions import db
from app.models import Product

BARCODE_HEADERS = {"баркод", "barcode", "штрихкод", "штрих-код"}
PRICE_HEADERS = {"цена закуп", "цена закуп.", "закупочная цена", "purchase price", "purchase_price"}

GOOGLE_SHEETS_RE = re.compile(
    r"https?://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9-_]+)",
    re.IGNORECASE,
)


def normalize_sheet_url(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    match = GOOGLE_SHEETS_RE.match(url.split("#")[0])
    if not match:
        return url
    sheet_id = match.group(1)
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    gid = (query.get("gid") or ["0"])[0]
    return (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export"
        f"?format=xlsx&gid={gid}"
    )


def normalize_barcode(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, float):
        if value != value:
            return None
        if value == int(value):
            text = str(int(value))
        else:
            text = str(value).strip()
    elif isinstance(value, int):
        text = str(value)
    else:
        text = str(value).strip()
    if not text or text.lower() in {"none", "nan"}:
        return None
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _normalize_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _parse_price(value) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:
            return None
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "").replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _find_columns(headers: list) -> tuple[int | None, int | None]:
    barcode_col = None
    price_col = None
    for idx, header in enumerate(headers):
        norm = _normalize_header(header)
        if norm in BARCODE_HEADERS:
            barcode_col = idx
        if norm in PRICE_HEADERS or norm.startswith("цена закуп"):
            price_col = idx
    return barcode_col, price_col


def _parse_workbook(content: bytes) -> dict[str, Decimal]:
    if content[:2] == b"PK":
        return _parse_xlsx(content)
    return _parse_xls(content)


def _parse_xlsx(content: bytes) -> dict[str, Decimal]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _rows_to_prices(rows)


def _parse_xls(content: bytes) -> dict[str, Decimal]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(r) for r in range(sheet.nrows)]
    return _rows_to_prices(rows)


def _rows_to_prices(rows: list) -> dict[str, Decimal]:
    if not rows:
        raise ValueError("Файл пуст.")

    header_row = None
    header_index = 0
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        bc, pr = _find_columns(list(row))
        if bc is not None and pr is not None:
            header_row = list(row)
            header_index = i
            break

    if header_row is None:
        raise ValueError('Не найдены колонки «Баркод» и «Цена закуп».')

    barcode_col, price_col = _find_columns(header_row)
    prices: dict[str, Decimal] = {}

    for row in rows[header_index + 1 :]:
        if not row or len(row) <= max(barcode_col, price_col):
            continue
        barcode = normalize_barcode(row[barcode_col])
        price = _parse_price(row[price_col])
        if barcode and price is not None:
            prices[barcode] = price

    if not prices:
        raise ValueError("В файле нет строк с баркодом и ценой.")
    return prices


def fetch_purchase_prices_map(url: str) -> dict[str, Decimal]:
    download_url = normalize_sheet_url(url)
    try:
        resp = requests.get(download_url, timeout=60)
    except requests.RequestException as exc:
        raise RuntimeError(f"Не удалось скачать файл: {exc}") from exc
    if resp.status_code != 200:
        raise RuntimeError(f"Ошибка загрузки файла: HTTP {resp.status_code}")
    return _parse_workbook(resp.content)


def apply_purchase_prices(user) -> dict:
    url = (user.purchase_prices_url or "").strip()
    if not url:
        return {"ok": True, "skipped": True, "updated": 0, "message": None}

    try:
        price_map = fetch_purchase_prices_map(url)
    except Exception as exc:
        return {"ok": False, "error": str(exc), "updated": 0}

    products = Product.query.filter_by(user_id=user.id).all()
    updated = 0
    for product in products:
        barcode = normalize_barcode(product.barcode)
        if not barcode:
            continue
        price = price_map.get(barcode)
        if price is None:
            continue
        product.purchase_price = price
        updated += 1

    db.session.flush()
    return {
        "ok": True,
        "skipped": False,
        "updated": updated,
        "total_in_file": len(price_map),
        "message": f"Закупочные цены обновлены: {updated} из {len(price_map)} в файле.",
    }
